import csv
import json
import re
from dataclasses import dataclass
from io import BytesIO, StringIO


@dataclass
class ParsedSection:
    text: str
    section_title: str = ""
    content_type: str = "text"
    page: int | None = None
    sheet: str = ""
    row_start: int | None = None
    row_end: int | None = None


def decode_bytes(file_bytes):
    for encoding in ["utf-8", "utf-8-sig", "gb18030"]:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


def sections_from_plain_text(text):
    sections = []
    current_title = ""
    current_lines = []

    for line in text.splitlines():
        stripped = line.strip()
        heading_match = re.match(r"^(#{1,6}\s+.+|[\d一二三四五六七八九十]+[、.]\s*.+)$", stripped)
        if heading_match and current_lines:
            sections.append(ParsedSection(
                text="\n".join(current_lines).strip(),
                section_title=current_title,
                content_type="text",
            ))
            current_lines = []

        if heading_match:
            current_title = stripped.lstrip("#").strip()
            current_lines.append(stripped)
        elif stripped:
            current_lines.append(stripped)

    if current_lines:
        sections.append(ParsedSection(
            text="\n".join(current_lines).strip(),
            section_title=current_title,
            content_type="text",
        ))

    return sections


def read_pdf_sections(file_bytes):
    import pypdf

    reader = pypdf.PdfReader(BytesIO(file_bytes))
    sections = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = page.extract_text()
        if text:
            sections.append(ParsedSection(
                text=text,
                section_title=f"第 {page_index} 页",
                content_type="pdf_page",
                page=page_index,
            ))
    return sections


def read_docx_sections(file_bytes):
    from docx import Document

    doc = Document(BytesIO(file_bytes))
    sections = []
    current_title = ""
    current_lines = []

    for paragraph in doc.paragraphs:
        text = paragraph.text.strip()
        if not text:
            continue

        style_name = paragraph.style.name.lower() if paragraph.style else ""
        is_heading = "heading" in style_name or "标题" in style_name

        if is_heading and current_lines:
            sections.append(ParsedSection(
                text="\n".join(current_lines).strip(),
                section_title=current_title,
                content_type="docx_section",
            ))
            current_lines = []

        if is_heading:
            current_title = text

        current_lines.append(text)

    if current_lines:
        sections.append(ParsedSection(
            text="\n".join(current_lines).strip(),
            section_title=current_title,
            content_type="docx_section",
        ))

    return sections


def read_csv_sections(file_bytes, rows_per_section=30):
    text = decode_bytes(file_bytes)
    rows = list(csv.reader(StringIO(text)))
    sections = []

    for start in range(0, len(rows), rows_per_section):
        batch = rows[start:start + rows_per_section]
        content = "\n".join(" | ".join(row) for row in batch)
        sections.append(ParsedSection(
            text=content,
            section_title=f"CSV 行 {start + 1}-{start + len(batch)}",
            content_type="table",
            row_start=start + 1,
            row_end=start + len(batch),
        ))

    return sections


def read_xlsx_sections(file_bytes, rows_per_section=30):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sections = []
    for sheet in workbook.worksheets:
        current_rows = []
        row_start = 1
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                current_rows.append(" | ".join(values))

            if len(current_rows) >= rows_per_section:
                sections.append(ParsedSection(
                    text="\n".join(current_rows),
                    section_title=f"{sheet.title} 行 {row_start}-{row_start + len(current_rows) - 1}",
                    content_type="table",
                    sheet=sheet.title,
                    row_start=row_start,
                    row_end=row_start + len(current_rows) - 1,
                ))
                row_start += len(current_rows)
                current_rows = []

        if current_rows:
            sections.append(ParsedSection(
                text="\n".join(current_rows),
                section_title=f"{sheet.title} 行 {row_start}-{row_start + len(current_rows) - 1}",
                content_type="table",
                sheet=sheet.title,
                row_start=row_start,
                row_end=row_start + len(current_rows) - 1,
            ))

    return sections


def read_json_sections(file_bytes):
    text = decode_bytes(file_bytes)
    data = json.loads(text)
    if isinstance(data, list):
        return [
            ParsedSection(
                text=json.dumps(item, ensure_ascii=False, indent=2),
                section_title=f"JSON item {index + 1}",
                content_type="json",
                row_start=index + 1,
            )
            for index, item in enumerate(data)
        ]

    if isinstance(data, dict):
        return [
            ParsedSection(
                text=json.dumps(value, ensure_ascii=False, indent=2),
                section_title=str(key),
                content_type="json",
            )
            for key, value in data.items()
        ]

    return [ParsedSection(
        text=json.dumps(data, ensure_ascii=False, indent=2),
        section_title="JSON",
        content_type="json",
    )]


def read_upload_as_sections(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    file_name = uploaded_file.name.lower()

    if file_name.endswith((".txt", ".md", ".log")):
        return sections_from_plain_text(decode_bytes(file_bytes))
    if file_name.endswith(".pdf"):
        return read_pdf_sections(file_bytes)
    if file_name.endswith(".docx"):
        return read_docx_sections(file_bytes)
    if file_name.endswith(".csv"):
        return read_csv_sections(file_bytes)
    if file_name.endswith(".xlsx"):
        return read_xlsx_sections(file_bytes)
    if file_name.endswith(".json"):
        return read_json_sections(file_bytes)

    raise ValueError("暂不支持这个文件格式")
