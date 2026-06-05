import csv
import json
import os
from io import BytesIO, StringIO

import streamlit as st


def get_secret(name):
    try:
        value = st.secrets[name]
    except Exception:
        value = os.getenv(name, "")
    return value


deepseek_key = get_secret("DEEPSEEK_API_KEY")
dashscope_key = get_secret("DASHSCOPE_API_KEY")

if deepseek_key:
    os.environ["DEEPSEEK_API_KEY"] = deepseek_key
if dashscope_key:
    os.environ["DASHSCOPE_API_KEY"] = dashscope_key

import rag_agent_core as agent

agent.seed_local_note()


st.set_page_config(
    page_title="RAG Agent Pro",
    page_icon="🤖",
    layout="wide",
)

st.title("RAG Agent Pro")


def decode_bytes(file_bytes):
    for encoding in ["utf-8", "utf-8-sig", "gb18030"]:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="ignore")


def read_pdf(file_bytes):
    import pypdf

    reader = pypdf.PdfReader(BytesIO(file_bytes))
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def read_docx(file_bytes):
    from docx import Document

    doc = Document(BytesIO(file_bytes))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def read_csv(file_bytes):
    text = decode_bytes(file_bytes)
    rows = csv.reader(StringIO(text))
    return "\n".join(" | ".join(row) for row in rows)


def read_xlsx(file_bytes):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []
    for sheet in workbook.worksheets:
        parts.append(f"工作表：{sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                parts.append(" | ".join(values))
    return "\n".join(parts)


def read_json(file_bytes):
    text = decode_bytes(file_bytes)
    data = json.loads(text)
    return json.dumps(data, ensure_ascii=False, indent=2)


def read_upload_as_text(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    file_name = uploaded_file.name.lower()

    if file_name.endswith((".txt", ".md", ".log")):
        return decode_bytes(file_bytes)
    if file_name.endswith(".pdf"):
        return read_pdf(file_bytes)
    if file_name.endswith(".docx"):
        return read_docx(file_bytes)
    if file_name.endswith(".csv"):
        return read_csv(file_bytes)
    if file_name.endswith(".xlsx"):
        return read_xlsx(file_bytes)
    if file_name.endswith(".json"):
        return read_json(file_bytes)

    raise ValueError("暂不支持这个文件格式")


def is_image(uploaded_file):
    return uploaded_file.type.startswith("image/")


def file_key(uploaded_file):
    return f"{uploaded_file.name}:{len(uploaded_file.getvalue())}"


def ingest_uploaded_files(uploaded_files, question):
    if not uploaded_files:
        return []

    ingested_sources = []

    for uploaded_file in uploaded_files:
        key = file_key(uploaded_file)
        if key in st.session_state.ingested_uploads:
            ingested_sources.append(st.session_state.ingested_uploads[key])
            continue

        if is_image(uploaded_file):
            if not dashscope_key:
                st.warning(f"{uploaded_file.name} 是图片，需要配置 DASHSCOPE_API_KEY 才能解析。")
                continue

            summary = agent.describe_image_bytes(
                uploaded_file.getvalue(),
                uploaded_file.type,
                "请提取这张图片中的关键信息，整理成适合知识库检索的文字资料。",
            )
            source = f"图片：{uploaded_file.name}"
            chunk_count = agent.add_text_to_chroma(
                summary,
                source=source,
                source_type="upload",
                url=uploaded_file.name,
            )
        else:
            text = read_upload_as_text(uploaded_file)
            source = f"上传：{uploaded_file.name}"
            chunk_count = agent.add_text_to_chroma(
                text,
                source=source,
                source_type="upload",
                url=uploaded_file.name,
            )

        st.session_state.ingested_uploads[key] = source
        st.session_state.upload_status.append(f"{source}：{chunk_count} 块")
        ingested_sources.append(source)

    return ingested_sources


def source_label(source):
    source_type = source.get("source_type", "unknown")
    title = source.get("source", "")

    if source_type == "upload" and title.startswith("图片："):
        return "上传图片｜优先"
    if source_type == "upload":
        return "上传资料｜优先"
    if source_type == "web":
        return "网络资料｜补充"
    if source_type == "local":
        return "基础资料｜兜底"
    return "其他资料｜参考"


with st.sidebar:
    st.header("资料")
    uploaded_files = st.file_uploader(
        "上传文件或图片",
        type=[
            "txt",
            "md",
            "log",
            "pdf",
            "docx",
            "csv",
            "xlsx",
            "json",
            "png",
            "jpg",
            "jpeg",
            "webp",
        ],
        accept_multiple_files=True,
    )

    top_k = st.slider("资料条数", 1, 5, 3)
    web_max_results = st.slider("网页结果数", 1, 5, 2)

    st.divider()
    st.caption("Agent 会自动使用上传资料，并联网补充资料；没有上传资料时，会直接联网收集。")
    st.write("DeepSeek:", "已配置" if deepseek_key else "未配置")
    st.write("通义百炼:", "已配置" if dashscope_key else "未配置")

    if "upload_status" in st.session_state and st.session_state.upload_status:
        st.divider()
        st.caption("已入库资料")
        for item in st.session_state.upload_status[-8:]:
            st.write(item)


if "messages" not in st.session_state:
    st.session_state.messages = []

if "last_sources" not in st.session_state:
    st.session_state.last_sources = []

if "ingested_uploads" not in st.session_state:
    st.session_state.ingested_uploads = {}

if "upload_status" not in st.session_state:
    st.session_state.upload_status = []


for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.write(message["content"])


prompt = st.chat_input("输入问题，Agent 会自动检索上传资料和网络资料")

if prompt:
    if not deepseek_key:
        st.error("请先配置 DEEPSEEK_API_KEY。")
        st.stop()

    st.session_state.messages.append({"role": "user", "content": prompt})

    with st.chat_message("user"):
        st.write(prompt)

    with st.chat_message("assistant"):
        try:
            with st.spinner("整理资料、联网检索、生成回答中..."):
                uploaded_sources = ingest_uploaded_files(uploaded_files, prompt)

                result = agent.answer_with_rag(
                    prompt,
                    use_web=True,
                    top_k=top_k,
                    web_max_results=web_max_results,
                    preferred_sources=uploaded_sources,
                )

            st.write(result["answer"])
            st.session_state.messages.append({
                "role": "assistant",
                "content": result["answer"],
            })
            st.session_state.last_sources = result["sources"]

            with st.expander("查看参考来源"):
                for index, source in enumerate(result["sources"], start=1):
                    title = source["source"]
                    url = source.get("url", "")
                    source_type = source.get("source_type", "unknown")
                    label = source_label(source)
                    st.markdown(f"**{index}. {title}**")
                    st.markdown(f"`{label}`")
                    st.caption(f"类型：{source_type}")
                    if url:
                        st.write(url)
                    st.write(source["document"][:300])
                    st.divider()

        except Exception as e:
            st.error(f"调用失败：{e}")
